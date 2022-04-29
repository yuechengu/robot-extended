import openpyxl
import os

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side


# 二维列表写入excel
def writeToExcel(title_list, table_list, file_name):
    # 存放位置
    file_path = os.path.abspath(os.path.dirname(__file__)) + fr"\..\data\{file_name}.xlsx"
    # workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = file_name
    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    # 标题行
    for t in range(len(title_list)):
        # 格式设置
        ws.cell(1, t + 1).font = Font(name=u'微软雅黑', bold=True)
        ws.cell(1, t + 1).fill = PatternFill('solid', fgColor='ffff00')
        ws.cell(1, t + 1).border = border
        # 写入
        ws.cell(1, t + 1).value = title_list[t]
    # 内容行
    for r in range(len(table_list)):
        for c in range(len(table_list[0])):
            # 格式设置
            ws.cell(r + 2, c + 1).font = Font(name=u'微软雅黑')  # excel中的行和列是从1开始计数的，需要+1
            ws.cell(r + 2, c + 1).border = border
            # 写入
            ws.cell(r + 2, c + 1).value = table_list[r][c]
    # 写入后保存
    wb.save(file_path)
    print("成功写入文件: " + file_path)
    os.startfile(file_path)
    return 0


# 从excel读取二维列表
def readFromExcel(file_name):
    # 存放位置
    file_path = os.path.abspath(os.path.dirname(__file__)) + fr"\..\data\{file_name}.xlsx"
    wb = load_workbook(file_path)
    ws = wb.active

    real_row_number = 0
    for i in range(0, ws.max_row):
        if ws.cell(i + 2, 1).value is None:
            real_row_number = i
            break

    table_list = [[] for i in range(real_row_number)]
    for i in range(0, real_row_number):
        table_list[i].append(ws.cell(i + 2, 1).value)
        table_list[i].append(ws.cell(i + 2, 2).value)
        table_list[i].append(ws.cell(i + 2, 3).value)
        table_list[i].append(ws.cell(i + 2, 4).value)

    return table_list


# 测试脚本
if __name__ == '__main__':
    # writeToExcel测试用例
    test_titles = ['A', 'B', 'C', 'D', 'E']
    test_table = [[1, 2, 3, 4, 5], [6, 7, 8, 9, 10], [11, 12, 13, 14, 15]]
    print("运行开始")
    writeToExcel(test_titles, test_table, 'test')

    # readFromExcel测试用例
    print(readFromExcel('SalesData'))
